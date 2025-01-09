from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
import re
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import time  # Biblioteca para controle de delays
from urllib.parse import quote_plus


# Função para verificar o tempo restante da licença de uso
def verificar_tempo_de_uso():
    return 30  # Exemplo: retorna 30 dias restantes

# Configuração do Selenium para usar o navegador Microsoft Edge
def configurar_driver():
    options = Options()
    options.add_argument("--remote-debugging-port=9222")  # Porta para debug remoto
    service = Service("C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe")  # Caminho do Edge WebDriver
    driver = webdriver.Edge(service=service, options=options)
    return driver

# Função que realiza o scraping dos dados do Google Maps
def capturar_dados_google_maps(driver):
    try:
        # Verifica o tempo de uso da licença antes de prosseguir
        dias_restantes = verificar_tempo_de_uso()
        if dias_restantes > 0:
            # Acessa a página do Google Maps
            driver.get("https://www.google.com/maps")
            
            # Aguarda o carregamento da página
            time.sleep(45)  # Espera 45 segundos antes de continuar

            # Localiza os elementos HTML necessários na página
            start_element = driver.find_element(By.XPATH, '//div[@jstcache="3"]')
            end_element = driver.find_element(By.XPATH, '//div[@jstcache="4"]')
            full_html = driver.page_source
        else:
            raise Exception("Licença Expirada")

        # Extrai o conteúdo da página entre os elementos encontrados
        start_html = start_element.get_attribute('outerHTML')
        end_html = end_element.get_attribute('outerHTML')
        start_index = full_html.find(start_html)
        end_index = full_html.find(end_html) + len(end_html)
        extracted_html = full_html[start_index:end_index]

        # Salva o conteúdo extraído em um arquivo TXT
        output_file = "output.txt"
        with open(output_file, "w", encoding="utf-8") as file:
            file.write(extracted_html)

        driver.quit()

        # Processa os dados extraídos do arquivo TXT
        with open(output_file, 'r', encoding='utf-8') as arquivo:
            conteudo = arquivo.read()

        # Expressões regulares para buscar os dados de nome, telefone e endereço
        padrao_nome = r'class="hfpxzc" aria-label="([^"]+)"'
        padrao_telefone = r'class="UsdlK">([^<]+)<'
        padrao_endereco = r'href="([^<]+)"'
        padrao_nome_arquivo = r'="Resultados para([^<]+)" role="feed" tabindex='

        # Realiza a extração dos dados com base nas expressões regulares
        nomes = re.finditer(padrao_nome, conteudo)
        telefones = re.finditer(padrao_telefone, conteudo)
        enderecos = re.finditer(padrao_endereco, conteudo)
        nome_arquivo = re.search(padrao_nome_arquivo, conteudo)

        # Determina o nome do arquivo a partir do conteúdo extraído
        if nome_arquivo:
            nome_arquivo = nome_arquivo.group(1).replace(', ', '-')  # Substitui vírgulas por traços
        else:
            nome_arquivo = "dados"  # Nome padrão caso o padrão não seja encontrado


        # Caminho da pasta de Downloads padrão do usuário
        caminho_downloads = os.path.join(os.path.expanduser('~'), 'Downloads')

        # Definindo os caminhos para salvar os arquivos gerados na pasta de Downloads
        caminho_saida_excel = os.path.join(caminho_downloads, f"{nome_arquivo}_banco_de_dados.xlsx")
        caminho_saida_csv = os.path.join(caminho_downloads, f"{nome_arquivo}_banco_de_dados.csv")


        # Organiza os dados extraídos em uma lista para formar um DataFrame
        dados = []
        iter_nomes = iter(nomes)
        iter_telefones = iter(telefones)
        iter_enderecos = iter(enderecos)

        nome_atual = next(iter_nomes, None)
        telefone_atual = next(iter_telefones, None)
        endereco_atual = next(iter_enderecos, None)

        # Preenche a lista de dados com os valores extraídos
        while nome_atual or telefone_atual or endereco_atual:
            if nome_atual and (not telefone_atual or nome_atual.start() < telefone_atual.start()) and (not endereco_atual or nome_atual.start() < endereco_atual.start()):
                dados.append([nome_atual.group(1), "", "", ""])  # Nome
                nome_atual = next(iter_nomes, None)
            elif telefone_atual and (not endereco_atual or telefone_atual.start() < endereco_atual.start()):
                if dados:
                    dados[-1][1] = telefone_atual.group(1)  # Telefone
                telefone_atual = next(iter_telefones, None)
            elif endereco_atual:
                if dados:
                    if dados[-1][2] == "":
                        dados[-1][2] = endereco_atual.group(1)  # Primeiro endereço
                    elif dados[-1][3] == "":
                        dados[-1][3] = endereco_atual.group(1)  # Segundo endereço
                endereco_atual = next(iter_enderecos, None)

        # Cria um DataFrame com os dados extraídos
        df = pd.DataFrame(dados, columns=["Nome", "Telefone", "Endereço", "Endereço 2"])
        
        # Salva os dados em arquivos Excel e CSV
        df.to_excel(caminho_saida_excel, index=False)
        df.to_csv(caminho_saida_csv, index=False, sep=";", encoding="utf-8")

        # Ajusta o layout da planilha Excel
        wb = load_workbook(caminho_saida_excel)
        ws = wb.active
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 60

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=12)

        wb.save(caminho_saida_excel)

        return caminho_saida_excel, caminho_saida_csv

    except Exception as e:
        raise Exception(f"Ocorreu um erro: {e}")
